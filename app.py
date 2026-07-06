import streamlit as st
import pandas as pd
import openpyxl
from ortools.sat.python import cp_model
import unicodedata
import re
import os

# テンプレートファイルのパス（xlsmを優先、なければxlsx）
if os.path.exists("template.xlsm"):
    TEMPLATE_FILE_PATH = "template.xlsm"
    TEMPLATE_EXT = ".xlsm"
else:
    TEMPLATE_FILE_PATH = "template.xlsx"
    TEMPLATE_EXT = ".xlsx"

def normalize_val(val):
    if pd.isna(val):
        return ""
    return unicodedata.normalize('NFKC', str(val)).strip()

def parse_classes(class_str):
    if pd.isna(class_str):
        return []
    parts = re.split(r'[\n,、\s]+', str(class_str))
    return [p.strip() for p in parts if p.strip()]

def parse_cell_address(cell_str):
    match = re.match(r"^([A-Za-z]+)(\d+)$", str(cell_str).strip())
    if not match:
        return None, None
    col_str = match.group(1).upper()
    row_str = match.group(2)
    col_num = 0
    for char in col_str:
        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    return int(row_str), col_num

import itertools

# ==========================================
# 開発者用設定（固定テンプレート・マッピング）
# ==========================================
# 使用するテンプレートファイルの名前を固定します
TEMPLATE_FILE_PATH = "template.xlsx"

# 曜日・時限の書き込みを開始する列番号（例: A列=1, B列=2, C列=3）
START_COL_INDEX = 5

# クラスがテンプレートの「何行目」に書き込まれるかの固定マッピング
# 左側がデータ上の名前（例: "1-1"）、右側が書き込む行番号
CLASS_ROW_MAPPING = {
    # クラスの行番号（実際の行番号に合わせて修正してください）
    "1-1": 5,
    "1-2": 6,
    "1-3": 7,
    "1-4": 8,
    "1-5": 9,
    "2-1": 10,
    "2-2": 11,
    "2-3": 12,
    "2-4": 13,
    "2-5": 14,
    "3-1": 15,
    "3-2": 16,
    "3-3": 17,
    "3-4": 18,
    "3-5": 19,
}
# ==========================================

def get_tt_leader_map(df, teacher_col, class_col, tt_col):
    row_classes_map = {}
    for idx, row in df.iterrows():
        row_classes_map[idx] = parse_classes(row[class_col])
        
    tt_dict = {}
    for idx, row in df.iterrows():
        t = row[teacher_col]
        if pd.isna(t): continue
        
        if tt_col and tt_col in df.columns:
            val = str(row[tt_col]).strip()
            if val and val != 'nan':
                if val not in tt_dict:
                    tt_dict[val] = []
                tt_dict[val].append(idx)
                
    tt_groups = {}
    tt_leader_map = {}
    for val, group in tt_dict.items():
        if len(group) > 1:
            leader = group[0]
            tt_groups[leader] = group
            for follower in group:
                tt_leader_map[follower] = leader
                
    return tt_leader_map, row_classes_map, tt_groups

def check_feasibility_diagnostics(df, teacher_col, class_col, hours_col, timeslot_cols, tt_col, block_class_on_fixed):
    diagnostics = []
    tt_leader_map, row_classes_map, tt_groups = get_tt_leader_map(df, teacher_col, class_col, tt_col)
    
    # 1. Row check (including teacher's global fixed slots)
    for idx, row in df.iterrows():
        t = row[teacher_col]
        if pd.isna(t): continue
        try:
            hours = int(row[hours_col])
        except:
            hours = 0
            
        t_rows = df[df[teacher_col] == t].index.tolist()
        available_slots = 0
        for col in timeslot_cols:
            val_str = normalize_val(row[col])
            if val_str in ['1', '1.0', '〇', '○', '◯']:
                # check if blocked by teacher's other fixed slots
                is_blocked = False
                for other_idx in t_rows:
                    other_val = normalize_val(df.loc[other_idx, col])
                    if other_val != '' and other_val not in ['1', '1.0', '〇', '○', '◯']:
                        is_blocked = True
                        break
                if not is_blocked:
                    available_slots += 1
                
        if hours > available_slots:
            cls_str = str(row[class_col])
            diagnostics.append(f"【{t}先生 - {cls_str}】必要な時数({hours}コマ)ですが、先生自身の他の固定コマと被っているため実際の空き枠が {available_slots} コマしかありません。")

    # 2. Class check
    classes = sorted(list(set([c for c_list in row_classes_map.values() for c in c_list])))
    for c in classes:
        c_rows = [idx for idx, c_list in row_classes_map.items() if c in c_list]
        unique_activity_rows = set()
        for idx in c_rows:
            unique_activity_rows.add(tt_leader_map.get(idx, idx))
            
        total_hours = sum([int(df.loc[idx, hours_col]) if pd.notna(df.loc[idx, hours_col]) and str(df.loc[idx, hours_col]).isdigit() else 0 for idx in unique_activity_rows])
        
        fixed_slots = 0
        if block_class_on_fixed:
            for col in timeslot_cols:
                has_fixed = False
                for idx in c_rows:
                    val_str = normalize_val(df.loc[idx, col])
                    if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                        has_fixed = True
                        break
                if has_fixed:
                    fixed_slots += 1
                
        max_avail = len(timeslot_cols) - fixed_slots
        if total_hours > max_avail:
            diagnostics.append(f"【クラス {c}】必要な授業数({total_hours}コマ)が、空き枠({max_avail}コマ)を超えています。")

    # 3. Teacher check
    teachers = df[teacher_col].dropna().unique()
    for t in teachers:
        t_rows = df[df[teacher_col] == t].index.tolist()
        total_hours = sum([int(df.loc[idx, hours_col]) if pd.notna(df.loc[idx, hours_col]) and str(df.loc[idx, hours_col]).isdigit() else 0 for idx in t_rows])
        
        teacher_available_slots = 0
        for col in timeslot_cols:
            has_fixed = False
            has_any_circle = False
            for idx in t_rows:
                val_str = normalize_val(df.loc[idx, col])
                if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                    has_fixed = True
                elif val_str in ['1', '1.0', '〇', '○', '◯']:
                    has_any_circle = True
            
            if has_any_circle and not has_fixed:
                teacher_available_slots += 1
                
        if total_hours > teacher_available_slots:
            diagnostics.append(f"【{t}先生】必要な全授業数({total_hours}コマ)に対して、実際の空き枠が {teacher_available_slots} コマしかありません。")

    # 4. TT check (local)
    for leader, group in tt_groups.items():
        if len(group) > 1:
            hours_set = set()
            for idx in group:
                try:
                    hours_set.add(int(df.loc[idx, hours_col]))
                except:
                    hours_set.add(0)
            
            if len(hours_set) > 1:
                teachers_in_tt = [str(df.loc[idx, teacher_col]) for idx in group]
                diagnostics.append(f"【TT設定矛盾】{'/'.join(teachers_in_tt)}先生のTTペアで、時数の設定が一致していません（{hours_set}）。TTは同じ時数にする必要があります。")
            else:
                h = list(hours_set)[0]
                if h > 0:
                    common_slots = 0
                    for col in timeslot_cols:
                        is_common = True
                        for idx in group:
                            val_str = normalize_val(df.loc[idx, col])
                            if val_str not in ['1', '1.0', '〇', '○', '◯']:
                                is_common = False
                                break
                            
                            t = df.loc[idx, teacher_col]
                            t_rows = df[df[teacher_col] == t].index.tolist()
                            for other_idx in t_rows:
                                other_val = normalize_val(df.loc[other_idx, col])
                                if other_val != '' and other_val not in ['1', '1.0', '〇', '○', '◯']:
                                    is_common = False
                                    break
                        if is_common:
                            common_slots += 1
                            
                    if h > common_slots:
                        teachers_in_tt = [str(df.loc[idx, teacher_col]) for idx in group]
                        diagnostics.append(f"【TT個別枠不足】{'/'.join(teachers_in_tt)}先生のペアは {h} コマ必要ですが、二人の予定が合う共通の空き枠が {common_slots} コマしかありません。")

    # 5. Global TT Pair check
    tt_pairs_hours = {}
    for leader, group in tt_groups.items():
        if len(group) > 1:
            teachers_in_tt = [str(df.loc[idx, teacher_col]) for idx in group]
            try:
                h = int(df.loc[group[0], hours_col])
            except:
                h = 0
                
            for t1, t2 in itertools.combinations(teachers_in_tt, 2):
                pair = tuple(sorted([t1, t2]))
                if pair not in tt_pairs_hours:
                    tt_pairs_hours[pair] = 0
                tt_pairs_hours[pair] += h
                
    for (t1, t2), required_shared_hours in tt_pairs_hours.items():
        t1_slots = set()
        t1_rows = df[df[teacher_col] == t1].index.tolist()
        for col in timeslot_cols:
            has_fixed = False
            has_circle = False
            for idx in t1_rows:
                val = normalize_val(df.loc[idx, col])
                if val != '' and val not in ['1', '1.0', '〇', '○', '◯']:
                    has_fixed = True
                elif val in ['1', '1.0', '〇', '○', '◯']:
                    has_circle = True
            if has_circle and not has_fixed:
                t1_slots.add(col)
                
        t2_slots = set()
        t2_rows = df[df[teacher_col] == t2].index.tolist()
        for col in timeslot_cols:
            has_fixed = False
            has_circle = False
            for idx in t2_rows:
                val = normalize_val(df.loc[idx, col])
                if val != '' and val not in ['1', '1.0', '〇', '○', '◯']:
                    has_fixed = True
                elif val in ['1', '1.0', '〇', '○', '◯']:
                    has_circle = True
            if has_circle and not has_fixed:
                t2_slots.add(col)
                
        common_slots = t1_slots.intersection(t2_slots)
        if required_shared_hours > len(common_slots):
            diagnostics.append(f"【全体TT枠不足】{t1}先生と{t2}先生は合計 {required_shared_hours} コマのTT授業を持っていますが、二人の予定が合う共通の空き枠が1週間全体で {len(common_slots)} コマしかありません。")

    return diagnostics

def export_to_excel_template(df_class, df_teacher, teacher_row_mapping, selected_week_str=None, start_date_str=None, end_date_str=None, teacher_subjects=None, nendo_str=None):
    from openpyxl import load_workbook
    from io import BytesIO
    import re
    import unicodedata
    import os
    
    if not os.path.exists(TEMPLATE_FILE_PATH):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません。同じフォルダに 'template.xlsx' または 'template.xlsm' を置いてください。")
        
    keep_vba = TEMPLATE_FILE_PATH.endswith('.xlsm')
    wb = load_workbook(TEMPLATE_FILE_PATH, keep_vba=keep_vba, keep_links=False)
    # 確実に1シート目を指定
    ws = wb.worksheets[0]
    
    if nendo_str:
        ws['A2'] = nendo_str
        
    if selected_week_str:
        ws.cell(row=1, column=40).value = selected_week_str  # AN1
        ws.cell(row=1, column=41).value = start_date_str     # AO1
        ws.cell(row=1, column=42).value = end_date_str       # AP1
    
    def norm_s(s):
        if pd.isna(s) or s is None: return ""
        return re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(s)))
        
    # クラスデータの準備
    class_data = {}
    for idx, row in df_class.iterrows():
        class_data[norm_s(idx)] = row.values
        
    # 教員データの準備
    teacher_data = {}
    for idx, row in df_teacher.iterrows():
        t_name = row['教員名']
        teacher_data[norm_s(t_name)] = row[2:].values 
        
    # クラスのマッピングをもとに書き込み
    for target_name, target_row in CLASS_ROW_MAPPING.items():
        c_norm = norm_s(target_name)
        if c_norm in class_data:
            data_to_write = class_data[c_norm]
            for i, val in enumerate(data_to_write):
                if pd.notna(val) and str(val).strip() != "":
                    ws.cell(row=target_row, column=START_COL_INDEX + i).value = str(val)
                    
    # 教員のマッピング（UIから取得）をもとに書き込み
    for t_name, target_row in teacher_row_mapping.items():
        c_norm = norm_s(t_name)
        if c_norm in teacher_data:
            ws.cell(row=target_row, column=3).value = t_name
            
            if teacher_subjects and t_name in teacher_subjects:
                subj = teacher_subjects[t_name]
                if subj:
                    ws.cell(row=target_row, column=4).value = subj[0]
                    
            data_to_write = teacher_data[c_norm]
            grade_class_counts = {'1': {}, '2': {}, '3': {}}
            
            for i, val in enumerate(data_to_write):
                if pd.notna(val) and str(val).strip() != "":
                    ws.cell(row=target_row, column=START_COL_INDEX + i).value = str(val)
                    for v in str(val).split('\n'):
                        v = v.strip()
                        if not v: continue
                        grade = v[0]
                        if grade in grade_class_counts:
                            grade_class_counts[grade][v] = grade_class_counts[grade].get(v, 0) + 1
                            
            ws.cell(row=target_row, column=56).value = max(grade_class_counts['1'].values()) if grade_class_counts['1'] else 0
            ws.cell(row=target_row, column=57).value = max(grade_class_counts['2'].values()) if grade_class_counts['2'] else 0
            ws.cell(row=target_row, column=58).value = max(grade_class_counts['3'].values()) if grade_class_counts['3'] else 0
                        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), TEMPLATE_EXT

def generate_timetable(df, teacher_col, class_col, hours_col, timeslot_cols, subject_cols, tt_col=None, lecturer_col=None, block_class_on_fixed=False, teacher_order=None, original_timeslot_cols=None, progress_bar=None, status_text=None):
    if original_timeslot_cols is None:
        original_timeslot_cols = timeslot_cols
        
    if progress_bar: progress_bar.progress(10)
    if status_text: status_text.text("データを読み込み中...")
        
    model = cp_model.CpModel()
    num_timeslots = len(timeslot_cols)
    assign = {}
    
    teachers = df[teacher_col].dropna().unique()
    tt_leader_map, row_classes_map, tt_groups = get_tt_leader_map(df, teacher_col, class_col, tt_col)
    
    individual_classes = sorted(list(set([c for c_list in row_classes_map.values() for c in c_list])))
    
    all_weighted_assign_vars = []
    
    tt_rows_set = set()
    for group in tt_groups.values():
        if len(group) > 1:
            for idx in group:
                tt_rows_set.add(idx)
                
    teacher_total_hours = {}
    for t in teachers:
        t_rows = df[df[teacher_col] == t].index.tolist()
        total_h = sum([int(df.loc[i, hours_col]) if pd.notna(df.loc[i, hours_col]) and str(df.loc[i, hours_col]).isdigit() else 0 for i in t_rows])
        teacher_total_hours[t] = total_h
    
    for idx, row in df.iterrows():
        t = row[teacher_col]
        if pd.isna(t):
            continue
            
        try:
            hours = int(row[hours_col])
        except:
            hours = 0
            
        # ベースの優先度スコア（桁を変えることで絶対に順位を逆転させない）
        base_weight = 1000
        if idx in tt_rows_set:
            base_weight = 10000
        if lecturer_col and lecturer_col in df.columns:
            val = normalize_val(row[lecturer_col])
            if val in ['1', '1.0', '〇', '○', '◯', 'True', 'true']:
                base_weight = 100000
                
        # サブ優先度：「週に入れる枠（持ちコマ数）」が少ない先生ほど優先（100 - 持ちコマ数）
        # ※最大でも持ちコマは40コマ程度を想定し、ベーススコアの桁を超えないようにする
        t_hours = teacher_total_hours.get(t, 0)
        sub_weight = max(0, 100 - t_hours) 
        
        weight = base_weight + sub_weight
            
        row_vars = []
        for p, col in enumerate(timeslot_cols):
            var = model.NewBoolVar(f'assign_{idx}_{p}')
            assign[(idx, p)] = var
            row_vars.append(var)
            
            import random
            slot_noise = random.randint(1, 50)
            slot_weight = (weight * 100) + slot_noise
            all_weighted_assign_vars.append(var * slot_weight)
            
            val_str = normalize_val(row[col])
            if val_str not in ['1', '1.0', '〇', '○', '◯']:
                model.Add(var == 0)
                
        if hours > 0:
            # 完璧に入らなくてもOKにするため、== ではなく <= に変更
            model.Add(sum(row_vars) <= hours)
        else:
            model.Add(sum(row_vars) == 0)

    # TT Sync constraints
    for leader, group in tt_groups.items():
        for follower in group:
            if follower != leader:
                for p in range(num_timeslots):
                    if (leader, p) in assign and (follower, p) in assign:
                        model.Add(assign[(leader, p)] == assign[(follower, p)])

    # Teacher Overlap
    for t in teachers:
        t_rows = df[df[teacher_col] == t].index.tolist()
        for p, col in enumerate(timeslot_cols):
            has_fixed = False
            for idx in t_rows:
                val_str = normalize_val(df.loc[idx, col])
                if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                    has_fixed = True
                    break
            
            vars_for_t_p = [assign[(idx, p)] for idx in t_rows if (idx, p) in assign]
            if has_fixed:
                if vars_for_t_p:
                    model.Add(sum(vars_for_t_p) == 0)
            else:
                if vars_for_t_p:
                    model.Add(sum(vars_for_t_p) <= 1)

    # Class Overlap
    for c in individual_classes:
        c_rows = [idx for idx, c_list in row_classes_map.items() if c in c_list]
        
        unique_activity_rows = set()
        for idx in c_rows:
            unique_activity_rows.add(tt_leader_map.get(idx, idx))
            
        for p, col in enumerate(timeslot_cols):
            has_fixed = False
            if block_class_on_fixed:
                for idx in c_rows:
                    val_str = normalize_val(df.loc[idx, col])
                    if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                        has_fixed = True
                        break
            
            vars_for_c_p = [assign[(idx, p)] for idx in unique_activity_rows if (idx, p) in assign]
            if has_fixed:
                if vars_for_c_p:
                    model.Add(sum(vars_for_c_p) == 0)
            else:
                if vars_for_c_p:
                    model.Add(sum(vars_for_c_p) <= 1)

    # 同じ教科が同じ日に複数回入らないようにする（1日の上限を設定）
    import math
    row_subject = {}
    for idx, row in df.iterrows():
        subj = "Unknown"
        if subject_cols:
            for s_col in subject_cols:
                val = str(row[s_col]).strip()
                if val != "" and val.lower() not in ["nan", "none"]:
                    subj = val
                    break
        row_subject[idx] = subj

    class_subject_rows = {}
    for c in individual_classes:
        c_rows = [idx for idx, c_list in row_classes_map.items() if c in c_list]
        for idx in c_rows:
            subj = row_subject[idx]
            key = (c, subj)
            if key not in class_subject_rows:
                class_subject_rows[key] = []
            leader_idx = tt_leader_map.get(idx, idx)
            if leader_idx not in class_subject_rows[key]:
                class_subject_rows[key].append(leader_idx)

    day_to_p = {}
    for p, col in enumerate(timeslot_cols):
        day = str(col)[0]
        if day not in day_to_p:
            day_to_p[day] = []
        day_to_p[day].append(p)
        
    num_days = len(day_to_p)
    for (c, subj), idx_list in class_subject_rows.items():
        if subj == "Unknown" or subj == "": continue
        
        total_hours = sum(int(df.loc[i, hours_col]) if pd.notna(df.loc[i, hours_col]) and str(df.loc[i, hours_col]).isdigit() else 0 for i in idx_list)
        
        max_per_day = 1
        if total_hours > num_days and num_days > 0:
            max_per_day = math.ceil(total_hours / num_days)
            
        for day, p_list in day_to_p.items():
            vars_for_day = []
            for p in p_list:
                for idx in idx_list:
                    if (idx, p) in assign:
                        vars_for_day.append(assign[(idx, p)])
            if vars_for_day:
                model.Add(sum(vars_for_day) <= max_per_day)

    # 全体として「優先度の高い授業からできるだけ多く配置する」ようAIに指示（最適化）
    model.Maximize(sum(all_weighted_assign_vars))

    if progress_bar: progress_bar.progress(50)
    if status_text: status_text.text("条件を満たす最適な組み合わせを探索しています...")
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    import random
    solver.parameters.random_seed = random.randint(1, 100000)
    status = solver.Solve(model)
        
    if progress_bar: progress_bar.progress(98)
    if status_text: status_text.text("結果を整理して表を作成中...")
    
    is_timeout = (status == cp_model.UNKNOWN)
    is_success = (status == cp_model.OPTIMAL or status == cp_model.FEASIBLE)
    
    if is_success:
        class_sched = {c: {col: "" for col in original_timeslot_cols} for c in individual_classes}
        teacher_sched = {t: {col: "" for col in original_timeslot_cols} for t in teachers}
        
        teacher_subjects = {t: set() for t in teachers}
        for idx, row in df.iterrows():
            t = row[teacher_col]
            if pd.isna(t): continue
            if subject_cols:
                for col in subject_cols:
                    if col in df.columns and pd.notna(row[col]) and str(row[col]).strip() and str(row[col]).strip() != "Unknown":
                        teacher_subjects[t].add(str(row[col]).strip())
                        break
                        
        for idx, row in df.iterrows():
            t = row[teacher_col]
            if pd.isna(t): continue
            c_list = row_classes_map.get(idx, [])
            for col in timeslot_cols:
                val_str = normalize_val(row[col])
                if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                    fixed_text = val_str
                    if block_class_on_fixed:
                        for c in c_list:
                            if not class_sched[c][col]:
                                class_sched[c][col] = fixed_text
                    if not teacher_sched[t][col]:
                        teacher_sched[t][col] = fixed_text
                        
        for (idx, p), var in assign.items():
            if solver.Value(var) == 1:
                col = timeslot_cols[p]
                t = df.loc[idx, teacher_col]
                c_list = row_classes_map.get(idx, [])
                
                c_display = "\n".join(c_list)
                if teacher_sched[t][col]:
                    if c_display not in teacher_sched[t][col]:
                        teacher_sched[t][col] += f" \n {c_display}"
                else:
                    teacher_sched[t][col] = c_display
                    
                subj = ""
                for s_col in subject_cols:
                    if pd.notna(df.loc[idx, s_col]) and str(df.loc[idx, s_col]).strip() and str(df.loc[idx, s_col]).strip() != "Unknown":
                        subj = str(df.loc[idx, s_col]).strip()
                        break
                display_text_c = subj if subj else f"{t}"
                for c in c_list:
                    if class_sched[c][col]:
                        if display_text_c not in class_sched[c][col]:
                            class_sched[c][col] += f" \n {display_text_c}"
                    else:
                        class_sched[c][col] = display_text_c
                        
        records = []
        class_hours = {c: 0 for c in individual_classes}
        for idx, row in df.iterrows():
            for c in row_classes_map.get(idx, []):
                class_hours[c] += int(row[hours_col]) if pd.notna(row[hours_col]) and str(row[hours_col]).isdigit() else 0
                
        for c in sorted(individual_classes):
            row = {"クラス": c}
            for col in original_timeslot_cols:
                row[col] = class_sched[c][col]
            records.append(row)
        df_class = pd.DataFrame(records).set_index("クラス")
        
        teacher_records = []
        sorted_teachers = sorted(teachers) if not teacher_order else ([t for t in teacher_order if t in teachers] + [t for t in teachers if t not in teacher_order])
        for t in sorted_teachers:
            row = {"教員名": t, "担当科目": ", ".join(sorted(list(teacher_subjects[t])))}
            for col in original_timeslot_cols:
                row[col] = teacher_sched[t][col]
            teacher_records.append(row)
        df_teacher = pd.DataFrame(teacher_records)
        
        unplaced_info = []
        for idx, row in df.iterrows():
            t = row[teacher_col]
            if pd.isna(t): continue
            try:
                hours = int(row[hours_col])
            except:
                hours = 0
            placed_count = sum(1 for p in range(num_timeslots) if (idx, p) in assign and solver.Value(assign[(idx, p)]) == 1)
            if placed_count < hours:
                cls_str = str(row[class_col])
                c_list = row_classes_map.get(idx, [])
                t_free_slots = []
                for p in range(num_timeslots):
                    has_fixed = False
                    t_rows = df[df[teacher_col] == t].index.tolist()
                    for other_idx in t_rows:
                        val_str = normalize_val(df.loc[other_idx, timeslot_cols[p]])
                        if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                            has_fixed = True
                            break
                    is_teaching = any(solver.Value(assign[(other_idx, p)]) == 1 for other_idx in t_rows if (other_idx, p) in assign)
                    if not has_fixed and not is_teaching:
                        t_free_slots.append(p)
                c_free_slots = []
                for p in range(num_timeslots):
                    c_available = True
                    for c in c_list:
                        c_rows = [i for i, clist in row_classes_map.items() if c in clist]
                        has_fixed = False
                        if block_class_on_fixed:
                            for c_idx in c_rows:
                                val_str = normalize_val(df.loc[c_idx, timeslot_cols[p]])
                                if val_str != '' and val_str not in ['1', '1.0', '〇', '○', '◯']:
                                    has_fixed = True
                                    break
                        is_taking = any(solver.Value(assign[(c_idx, p)]) == 1 for c_idx in c_rows if (c_idx, p) in assign)
                        if has_fixed or is_taking:
                            c_available = False
                            break
                    if c_available:
                        c_free_slots.append(p)
                reason = ""
                if len(t_free_slots) == 0:
                    reason = f"教員（{t}先生）の他の授業や固定予定で空き枠が埋まりきっているため"
                elif len(c_free_slots) == 0:
                    reason = f"クラス（{cls_str}）の時間割が他の授業で埋まりきっているため"
                else:
                    common = set(t_free_slots).intersection(set(c_free_slots))
                    if len(common) == 0:
                        reason = "教員とクラスの双方に空き枠はあるが、タイミングが合わなかったため"
                    else:
                        reason = "他の授業の配置との兼ね合い（玉突き）で配置枠が確保できなかったか、または全体として優先順位の調整がつかなかったため"
                unplaced_info.append(f"{t}先生の {cls_str} (目標 {hours}コマ中、{placed_count}コマ配置) ➡ **推測理由: {reason}**")

        return True, df_class, df_teacher, is_timeout, unplaced_info
    else:
        return False, None, None, is_timeout, []

@st.dialog("Excel表の開始位置を指定")
def start_cell_dialog():
    st.write("表が始まる左上のセル番地を入力してください。")
    cell = st.text_input("開始セル（例: B2）", value=st.session_state.get("start_cell", "B2"))
    if st.button("決定", use_container_width=True):
        st.session_state.start_cell = cell
        st.rerun()

@st.dialog("アップロードされた生データ", width="large")
def show_raw_data_dialog(df):
    st.dataframe(df, use_container_width=True)

@st.dialog("クラス別 時間割", width="large")
def show_class_timetable_dialog(df):
    st.dataframe(df, use_container_width=True)

@st.dialog("教員別 時間割", width="large")
def show_teacher_timetable_dialog(df):
    teacher_height = len(df) * 35 + 38
    st.dataframe(df, use_container_width=True, hide_index=True, height=teacher_height)

def main():
    st.set_page_config(page_title="時間割自動作成アプリ", layout="wide")
    
    st.title("時間割作成（テスト版）")
    st.markdown("""
        <div style="display: flex; align-items: center; margin-bottom: 1rem;">
            <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 48 48" width="32" height="32" style="margin-right: 12px; filter: drop-shadow(0px 2px 4px rgba(0,0,0,0.2));"><path fill="#4CAF50" d="M41,10H25v28h16c0.553,0,1-0.447,1-1V11C42,10.447,41.553,10,41,10z"/><path fill="#388E3C" d="M32 15H39V18H32zM32 25H39V28H32zM32 30H39V33H32zM32 20H39V23H32zM25 10H27V38H25z"/><path fill="#2E7D32" d="M27,42L6,38.203V9.797L27,6V42z"/><path fill="#FFF" d="M19.129,31l-2.411-4.561c-0.092-0.171-0.186-0.483-0.284-0.938h-0.037c-0.046,0.215-0.154,0.541-0.324,0.979L13.65,31H9.643l5.074-7.539L10.009,17h4.085l2.003,4.218c0.12,0.264,0.243,0.597,0.371,1.002h0.037c0.053-0.181,0.165-0.518,0.336-1.012L18.841,17h3.805l-4.521,6.866L23.111,31H19.129z"/></svg>
            <h3 style="margin: 0; padding: 0;">設定ファイルの読み込み</h3>
        </div>
    """, unsafe_allow_html=True)

    st.markdown("""
        <style>
        /* ファイルアップローダーをコンパクトなボタン型に変形するCSS */
        /* ドラッグ＆ドロップ領域の不要なテキストやアイコンを消すための安全なCSS */
        /* Dropzone全体に影響（文字を透明化してデフォルトテキストを消去） */
        [data-testid="stFileUploaderDropzone"] {
            color: transparent !important;
            margin: 0 !important;
            padding: 0 !important;
            border: none !important;
            background-color: transparent !important;
            min-height: auto !important;
            width: max-content !important; /* 余白がクリック可能になるのを防ぐため、幅をボタンに合わせる */
        }
        
        /* stFileUploader全体の余白を消去し、縦ズレを防止 */
        [data-testid="stFileUploader"] {
            margin: 0 !important;
            padding: 0 !important;
        }
        
        /* Dropzone内外の不要なテキスト（200MB制限など）やアイコンを安全に消去 */
        [data-testid="stFileUploaderDropzone"] svg,
        [data-testid="stFileUploaderDropzone"] [data-testid="stMarkdownContainer"],
        [data-testid="stFileUploader"] small,
        [data-testid="stFileUploaderInstructions"] {
            display: none !important;
        }
        
        /* アップロードボタン内の不要なアイコン等を消去 */
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"] > * {
            display: none !important;
        }
        
        /* ボタンのデザインを他の標準ボタン（データ確認、クリア）に完璧に合わせる */
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"] {
            font-size: 0 !important; /* 元の文字を幅ごと完全に消す */
            color: transparent !important;
            background-color: #ffffff !important;
            border: 1px solid rgba(49, 51, 63, 0.2) !important;
            border-radius: 8px !important;
            transition: all 0.2s ease !important;
            width: 100% !important; /* 幅いっぱいに広げる */
            min-height: 38px !important; /* 高さをデータ確認ボタンと完全に一致させる */
            padding: 0 16px !important; /* 左右の余白を適切に設定 */
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
        }
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"]:hover {
            border-color: #FF4B4B !important;
            color: transparent !important; /* hover時も透明を維持 */
        }
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"]:active {
            background-color: #FF4B4B !important;
            color: transparent !important;
        }
        /* 疑似要素でテキストを上書き（ボタンの寸法を自動拡張させるため絶対配置は使わない） */
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"]::after {
            content: "📁 ファイルをアップロード";
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 16px !important;
            font-weight: normal !important;
            color: #31333F !important;
        }
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"]:hover::after {
            color: #FF4B4B !important;
        }
        [data-testid="stFileUploaderDropzone"] button[data-testid="stBaseButton-secondary"]:active::after {
            color: #ffffff !important;
        }
        </style>
    """, unsafe_allow_html=True)

    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = 0
    if "saved_file" not in st.session_state:
        st.session_state.saved_file = None

    # アップロードボタンとデータ確認ボタンの隙間を最小限にするため、カラム幅を詰める
    col_upload, col_btn, col_clear, col_empty = st.columns([0.9, 0.7, 0.7, 2.5], vertical_alignment="bottom")
    with col_upload:
        # StreamlitのファイルリストUIを出さないため、アップロード直後にウィジェットを強制リセットする
        uploaded_files = st.file_uploader("設定ファイルの読み込み", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True, label_visibility="collapsed", key=f"uploader_{st.session_state.uploader_key}")
        
        if uploaded_files:
            file = uploaded_files[-1]
            import io
            class SavedFile(io.BytesIO):
                def __init__(self, name, size, data):
                    super().__init__(data)
                    self.name = name
                    self.size = size
            
            st.session_state.saved_file = SavedFile(file.name, file.size, file.getvalue())
            st.session_state.uploader_key += 1
            st.rerun()

    with col_btn:
        raw_data_btn_placeholder = st.empty()

    uploaded_file = st.session_state.saved_file

    with col_clear:
        if uploaded_file is not None:
            if st.button("✖ クリア", key="clear_file_btn", use_container_width=True):
                st.session_state.saved_file = None
                st.rerun()

    if uploaded_file is not None:
        # デフォルトUIを消した代わりに、テキストだけでファイル名とサイズを表示する
        excel_icon_svg = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 48 48" width="16px" height="16px" style="vertical-align: middle; margin-right: 5px; margin-top: -2px;"><path fill="#4CAF50" d="M41,10H25v28h16c0.553,0,1-0.447,1-1V11C42,10.447,41.553,10,41,10z"/><path fill="#FFF" d="M32,15h2v18h-2V15z"/><path fill="#4CAF50" d="M27.5,10.5l-15-3v33l15-3V10.5z"/><path fill="#FFF" d="M19.141,29.35l2.253-3.666l-2.029-3.957h2.646l1.085,2.449c0.165,0.373,0.279,0.73,0.34,1.071h0.046c0.076-0.34,0.198-0.697,0.366-1.071l1.157-2.449h2.464l-2.19,3.951l2.368,3.673h-2.617l-1.341-2.529c-0.18-0.354-0.334-0.704-0.463-1.05h-0.033c-0.129,0.34-0.276,0.684-0.44,1.034l-1.378,2.545H19.141z"/></svg>'''
        
        st.markdown(f"<div style='margin-top: 5px; margin-bottom: 20px; color: #444; font-size: 0.95rem; display: flex; align-items: center;'>{excel_icon_svg} <b>{uploaded_file.name}</b>&nbsp;({uploaded_file.size / 1024:.1f} KB)</div>", unsafe_allow_html=True)

        
        current_file_id = f"{uploaded_file.name}_{uploaded_file.size}"
        if st.session_state.get("last_uploaded_file") != current_file_id:
            st.session_state.last_uploaded_file = current_file_id
            
        start_cell = st.session_state.get("start_cell", "B2")
        
        with st.sidebar:
            st.markdown("### ⚙️ 詳細設定")
            test_mode_toggled = st.toggle("🔧 テストモード", key="test_mode_toggled", help="開発者用のデバッグ機能を有効にします")
            is_test_mode = False
            if test_mode_toggled:
                if not st.session_state.get("test_mode_unlocked", False):
                    pw = st.text_input("パスワードを入力", type="password")
                    if pw == "abc123":
                        st.session_state.test_mode_unlocked = True
                        st.rerun()
                    elif pw != "":
                        st.error("パスワードが違います")
                else:
                    is_test_mode = True
            else:
                st.session_state.test_mode_unlocked = False
            st.markdown("---")
            st.markdown("#### 設定ファイルの認識設定")
            st.caption("※通常は変更不要です")
            st.info(f"表の開始セル: **{start_cell}**")
            if st.button("開始セルを変更", use_container_width=True):
                start_cell_dialog()
                
        header_row, start_col = parse_cell_address(start_cell)

        if header_row is None or start_col is None:
            st.error("有効なセル番地（例: B2, A1 など）を入力してください。")
            return
            
        try:
            xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
            if len(xls.sheet_names) >= 2:
                # 1シート目（Index 0）を読み込み、D列（index 3）から教員順序、AB列（index 27）から行番号を取得
                df_order = pd.read_excel(xls, sheet_name=0, header=None, engine='openpyxl')
                teacher_order = None
                teacher_default_rows = {}
                if df_order.shape[1] >= 4:
                    teacher_order = []
                    for i in range(len(df_order)):
                        t_name = str(df_order.iloc[i, 3]).strip()
                        if t_name and t_name.lower() != "nan":
                            if t_name not in teacher_order:
                                teacher_order.append(t_name)
                            if df_order.shape[1] > 27:
                                val = df_order.iloc[i, 27]
                                if pd.notna(val) and str(val).strip() != "":
                                    try:
                                        teacher_default_rows[t_name] = int(float(str(val).strip()))
                                    except ValueError:
                                        pass
                st.session_state.teacher_default_rows = teacher_default_rows
                
                if len(xls.sheet_names) >= 3:
                    df_week = pd.read_excel(xls, sheet_name=2, header=None, engine='openpyxl')
                    week_options = []
                    week_data_map = {}
                    for i in range(1, len(df_week)):
                        row_vals = df_week.iloc[i, 1:4].dropna().tolist()
                        if len(row_vals) >= 3:
                            col1 = str(row_vals[0]).replace('.0', '')
                            def format_date(d):
                                if isinstance(d, pd.Timestamp):
                                    if d.year >= 2019:
                                        return f"R{d.year - 2018}.{d.month}.{d.day}"
                                    return f"{d.year}.{d.month}.{d.day}"
                                return str(d).split()[0]
                            col2 = format_date(row_vals[1])
                            col3 = format_date(row_vals[2])
                            week_str = f"【週案{col1}】　{col2} ～ {col3}"
                            week_options.append(week_str)
                            week_data_map[week_str] = {
                                'week': col1,
                                'start': row_vals[1],
                                'end': row_vals[2]
                            }
                    st.session_state.week_options = week_options
                    st.session_state.week_data_map = week_data_map
                else:
                    st.session_state.week_options = []
                    st.session_state.week_data_map = {}
                    
                # 2シート目（Index 1）をデータとして読み込み
                df = pd.read_excel(xls, sheet_name=1, header=header_row - 1, engine='openpyxl')
            else:
                st.session_state.week_options = []
                teacher_order = None
                df = pd.read_excel(xls, sheet_name=0, header=header_row - 1, engine='openpyxl')
                
            if start_col > 1:
                df = df.iloc[:, (start_col - 1):]
                
            st.success("ファイルの読み込みに成功しました！")
            if raw_data_btn_placeholder.button("📋データ確認", use_container_width=True):
                    show_raw_data_dialog(df)
                
            st.markdown("---")
            st.markdown("#### 1. 項目の設定")

            
            columns = df.columns.tolist()
            if len(columns) < 4:
                st.error("列数が不足しています。開始列の設定を見直してください。")
                return

            options_with_none = ["-- 指定なし --"] + columns
            
            def norm_col(s):
                import re
                return re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(s)))

            def get_idx(name, opts=options_with_none):
                if name in opts:
                    return opts.index(name)
                name_norm = norm_col(name)
                for i, opt in enumerate(opts):
                    if norm_col(opt) == name_norm:
                        return i
                # Fallback: substring match if exact normalized match fails
                for i, opt in enumerate(opts):
                    if name_norm in norm_col(opt):
                        return i
                return 0
                
            tt_options = ["-- 指定しない --"] + columns
            def get_tt_idx(name):
                return get_idx(name, tt_options)

            lec_options = ["-- 指定しない --"] + columns
            def get_lec_idx(name):
                return get_idx(name, lec_options)
            
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                teacher_col = st.selectbox("教員名", options_with_none, index=get_idx("名前"))
            with c2:
                class_col = st.selectbox("担当クラス", options_with_none, index=get_idx("担当"))
            with c3:
                target_subs = ["教科名(1年生)", "教科名(2年生)", "教科名(3年生)"]
                default_subjects = []
                for tgt in target_subs:
                    tgt_n = norm_col(tgt)
                    for col in columns:
                        if norm_col(col) == tgt_n or tgt_n in norm_col(col):
                            if col not in default_subjects:
                                default_subjects.append(col)
                            break
                subject_cols = st.multiselect("教科名（複数選択可）", columns, default=default_subjects)
            with c4:
                hours_col = st.selectbox("時数（週）", options_with_none, index=get_idx("時数(週)"))
                
            c5, c6, c7, c8 = st.columns(4)
            with c5:
                timeslot_start_col = st.selectbox("時間割（始）", options_with_none, index=get_idx("月1"))
            with c6:
                timeslot_end_col = st.selectbox("時間割（終）", options_with_none, index=get_idx("金6"))
            with c7:
                tt_col = st.selectbox("少人数ペア設定", tt_options, index=get_tt_idx("少人数ペア"), help="同じ文字（例: TT1）が入力された行同士を、必ず同じ時間に配置します。分担（別々の時間）の場合は指定しないでください。")
                if tt_col == "-- 指定しない --":
                    tt_col = None
            with c8:
                lecturer_col = st.selectbox("講師設定", lec_options, index=get_lec_idx("講師"), help="この列に1などの値が入っている行の授業を最優先で配置します。")
                if lecturer_col == "-- 指定しない --":
                    lecturer_col = None
            
            if "-- 指定なし --" in [teacher_col, class_col, hours_col, timeslot_start_col, timeslot_end_col]:
                st.warning("👆 項目の設定（教員名、クラス名、時数、始まる列、終わる列）をすべて選択してください。")
                return
                
            timeslot_start_idx = columns.index(timeslot_start_col)
            timeslot_end_idx = columns.index(timeslot_end_col)
            
            if timeslot_start_idx > timeslot_end_idx:
                st.error("時間割が始まる列は、終わる列より前である必要があります。")
                return
                
            original_timeslot_cols = columns[timeslot_start_idx:timeslot_end_idx + 1]
            
            st.markdown("---")
            st.markdown("#### 2. 授業枠の設定")
            
            day_map = {}
            for col in original_timeslot_cols:
                day = str(col)[0]
                if day not in day_map:
                    day_map[day] = []
                day_map[day].append(col)
                
            timeslot_cols = []
            
            if day_map:
                cols_layout = st.columns(len(day_map))
                for i, (day, slots) in enumerate(day_map.items()):
                    with cols_layout[i]:
                        st.markdown(f"**{day}曜日**")
                        for slot in slots:
                            if st.checkbox(slot, value=True, key=f"slot_{slot}"):
                                timeslot_cols.append(slot)
            else:
                timeslot_cols = original_timeslot_cols
            
            if not timeslot_cols:
                st.warning("有効な枠が1つも選択されていません。")
                return
            
            block_class_on_fixed = False

            target_filter = "すべての授業を作成（通常設定）"
            if is_test_mode:
                st.markdown("---")
                st.markdown("#### 4. 作成対象の絞り込み（テスト用）")
                target_filter = st.radio(
                    "どの授業を自動作成の対象にしますか？",
                    options=["すべての授業を作成（通常設定）", "講師の授業のみ作成（テスト用）", "少人数ペアの授業のみ作成（テスト用）", "それ以外の一般授業のみ作成（テスト用）"],
                    index=0
                )
                
                if "講師の授業" in target_filter and not lecturer_col:
                    st.warning("⚠️ 「講師設定」の列がマップされていないため、この条件では作成できません。")
                if "少人数ペア" in target_filter and not tt_col:
                    st.warning("⚠️ 「少人数ペア設定」の列がマップされていないため、この条件では作成できません。")

            st.markdown("---")
            if st.button("時間割を自動作成する", type="primary", use_container_width=True):
                if "講師の授業" in target_filter and not lecturer_col:
                    st.error("「講師設定」の列が指定されていません。")
                    return
                if "少人数ペア" in target_filter and not tt_col:
                    st.error("「少人数ペア設定」の列が指定されていません。")
                    return
                    
                progress_bar = st.progress(0)
                status_text = st.empty()
                with st.spinner("時間割作成の全プロセスを実行中...（最大30秒）"):
                    target_df = df.copy()
                    
                    def is_lecturer(v):
                        val = normalize_val(v)
                        return val in ['1', '1.0', '〇', '○', '◯', 'True', 'true']
                        
                    def is_tt(v):
                        val = normalize_val(v)
                        return val != "" and val.lower() not in ['0', '0.0', 'false', 'なし', 'nan', 'none']
                        
                    if "講師の授業" in target_filter and lecturer_col:
                        target_df = target_df[target_df[lecturer_col].apply(is_lecturer)]
                    elif "少人数ペア" in target_filter and tt_col:
                        target_df = target_df[target_df[tt_col].apply(is_tt)]
                    elif "それ以外" in target_filter:
                        if lecturer_col:
                            target_df = target_df[~target_df[lecturer_col].apply(is_lecturer)]
                        if tt_col:
                            target_df = target_df[~target_df[tt_col].apply(is_tt)]
                            
                    st.info(f"💡 抽出された対象授業データ: {len(target_df)}件 / 全{len(df)}件")
                            
                    if target_df.empty:
                        success = False
                        is_timeout = False
                        unplaced_info = []
                        st.session_state.diagnostics = ["対象となる授業データが見つかりませんでした。絞り込み条件とExcelデータを確認してください。"]
                    else:
                        success, df_class, df_teacher, is_timeout, unplaced_info = generate_timetable(
                            target_df, teacher_col, class_col, hours_col, timeslot_cols, subject_cols, tt_col, lecturer_col, block_class_on_fixed, teacher_order, original_timeslot_cols=original_timeslot_cols, progress_bar=progress_bar, status_text=status_text
                        )
                    
                if success:
                    progress_bar.progress(100)
                    status_text.text("✨ 時間割の作成が完了しました！")
                    import time
                    time.sleep(1)
                    progress_bar.empty()
                    status_text.empty()
                    st.session_state.success_status = "success"
                    st.session_state.df_class = df_class
                    st.session_state.df_teacher = df_teacher
                    st.session_state.unplaced_info = unplaced_info
                    if "teacher_mapping_df" in st.session_state:
                        del st.session_state.teacher_mapping_df
                else:
                    progress_bar.empty()
                    status_text.empty()
                    st.session_state.success_status = "failed"
                    st.session_state.is_timeout = is_timeout
                    st.session_state.diagnostics = check_feasibility_diagnostics(df, teacher_col, class_col, hours_col, timeslot_cols, tt_col, block_class_on_fixed)
                    
            if st.session_state.get("success_status") == "success":
                df_class = st.session_state.df_class
                df_teacher = st.session_state.df_teacher
                unplaced_info = st.session_state.unplaced_info
                
                if unplaced_info:
                    st.warning("⚠️ **一部の授業が配置できませんでした（できる分だけ限界まで配置しました）。以下の授業が未配置です:**")
                    for info in unplaced_info:
                        st.write(f"- {info}")
                    st.info("※未配置のコマは、手動で調整するか、空き枠（〇）を増やして再度自動作成をお試しください。")
                else:
                    st.success("🎉 すべての制約を満たす完璧な時間割が見つかりました！")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("クラス別時間割", use_container_width=True):
                        show_class_timetable_dialog(df_class)
                with col2:
                    if st.button("教員別時間割", use_container_width=True):
                        show_teacher_timetable_dialog(df_teacher)
                
                st.markdown("---")
                st.markdown("""
                    <div style="display: flex; align-items: center; margin-bottom: 1rem; margin-top: 1rem;">
                        <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 48 48" width="32" height="32" style="margin-right: 12px; filter: drop-shadow(0px 2px 4px rgba(0,0,0,0.2));"><path fill="#4CAF50" d="M41,10H25v28h16c0.553,0,1-0.447,1-1V11C42,10.447,41.553,10,41,10z"/><path fill="#388E3C" d="M32 15H39V18H32zM32 25H39V28H32zM32 30H39V33H32zM32 20H39V23H32zM25 10H27V38H25z"/><path fill="#2E7D32" d="M27,42L6,38.203V9.797L27,6V42z"/><path fill="#FFF" d="M19.129,31l-2.411-4.561c-0.092-0.171-0.186-0.483-0.284-0.938h-0.037c-0.046,0.215-0.154,0.541-0.324,0.979L13.65,31H9.643l5.074-7.539L10.009,17h4.085l2.003,4.218c0.12,0.264,0.243,0.597,0.371,1.002h0.037c0.053-0.181,0.165-0.518,0.336-1.012L18.841,17h3.805l-4.521,6.866L23.111,31H19.129z"/></svg>
                        <h3 style="margin: 0; padding: 0;">Excelテンプレートへ書き出し</h3>
                    </div>
                """, unsafe_allow_html=True)
                
                week_options = st.session_state.get("week_options", [])
                selected_week = None
                if week_options:
                    selected_week = st.selectbox("週の選択", options=["-- 出力しない --"] + week_options)
                    if selected_week == "-- 出力しない --":
                        selected_week = None
                
                teacher_names = df_teacher['教員名'].tolist()
                
                # セッションステート初期化（新しく生成されたときだけ実行される）
                if "teacher_mapping_df" not in st.session_state:
                    default_rows_dict = st.session_state.get("teacher_default_rows", {})
                    default_row_list = [default_rows_dict.get(t, None) for t in teacher_names]
                    st.session_state.teacher_mapping_df = pd.DataFrame({
                        "教員名": teacher_names,
                        "行番号": default_row_list
                    })
                
                # data_editorの出力を受け取るが、session_stateには上書きしない（Streamlitの仕様によるリセットを防ぐため）
                with st.expander("教員の出力先"):
                    edited_mapping_df = st.data_editor(st.session_state.teacher_mapping_df, use_container_width=True, hide_index=True)
                
                if st.button("テンプレートに書き込んで出力", type="primary", use_container_width=True):
                    try:
                        teacher_row_mapping = {}
                        for _, row in edited_mapping_df.iterrows():
                            if pd.notna(row['行番号']) and str(row['行番号']).strip() != "":
                                try:
                                    teacher_row_mapping[row['教員名']] = int(float(str(row['行番号']).strip()))
                                except ValueError:
                                    pass
                                    
                        teacher_subjects = {}
                        if subject_cols:
                            for idx, row in df.iterrows():
                                t = row[teacher_col]
                                if pd.isna(t): continue
                                t = str(t).strip()
                                if t not in teacher_subjects:
                                    for s_col in subject_cols:
                                        if pd.notna(row[s_col]):
                                            s_val = str(row[s_col]).strip()
                                            if s_val and s_val != "Unknown":
                                                teacher_subjects[t] = s_val
                                                break
                                                
                        week_str_parsed = None
                        start_date_parsed = None
                        end_date_parsed = None
                        nendo_str = None
                        if selected_week:
                            import re
                            m = re.search(r'【週案(.*?)】\s*(.*?)\s*～\s*(.*)', selected_week)
                            if m:
                                week_str_parsed = m.group(1)
                                start_date_parsed = m.group(2)
                                end_date_parsed = m.group(3)
                                
                            week_data_map = st.session_state.get("week_data_map", {})
                            if selected_week in week_data_map:
                                start_dt = week_data_map[selected_week]['start']
                                nendo_year = None
                                if isinstance(start_dt, pd.Timestamp):
                                    nendo_year = start_dt.year if start_dt.month >= 4 else start_dt.year - 1
                                elif isinstance(start_dt, str):
                                    start_str = start_dt.strip()
                                    if start_str.startswith("R") or start_str.startswith("r"):
                                        try:
                                            # R8.4.6
                                            parts = start_str[1:].split('.')
                                            r_year = int(parts[0])
                                            month = int(parts[1])
                                            calendar_year = r_year + 2018
                                            nendo_year = calendar_year if month >= 4 else calendar_year - 1
                                        except:
                                            pass
                                    else:
                                        try:
                                            parsed_dt = pd.to_datetime(start_dt)
                                            nendo_year = parsed_dt.year if parsed_dt.month >= 4 else parsed_dt.year - 1
                                        except:
                                            pass
                                
                                if nendo_year is not None:
                                    if nendo_year >= 2019:
                                        nendo_str = f"令和{nendo_year - 2018}年度"
                                    else:
                                        nendo_str = f"{nendo_year}年度"
                                
                        result_bytes, ext = export_to_excel_template(
                            df_class, df_teacher, teacher_row_mapping,
                            selected_week_str=week_str_parsed,
                            start_date_str=start_date_parsed,
                            end_date_str=end_date_parsed,
                            teacher_subjects=teacher_subjects,
                            nendo_str=nendo_str
                        )
                        
                        mime_type = "application/vnd.ms-excel.sheet.macroEnabled.12" if ext == ".xlsm" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        
                        out_filename = f"時間割_完成版{ext}"
                        if week_str_parsed:
                            out_filename = f"時間割_【週案{week_str_parsed}】{ext}"

                        st.success("書き込み準備が完了しました！下のボタンからダウンロードしてください。")
                        import base64
                        b64 = base64.b64encode(result_bytes).decode()
                        svg_icon = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 48 48" width="24" height="24" style="margin-right: 8px; filter: drop-shadow(0px 1px 2px rgba(0,0,0,0.2));"><path fill="#4CAF50" d="M41,10H25v28h16c0.553,0,1-0.447,1-1V11C42,10.447,41.553,10,41,10z"/><path fill="#388E3C" d="M32 15H39V18H32zM32 25H39V28H32zM32 30H39V33H32zM32 20H39V23H32zM25 10H27V38H25z"/><path fill="#2E7D32" d="M27,42L6,38.203V9.797L27,6V42z"/><path fill="#FFF" d="M19.129,31l-2.411-4.561c-0.092-0.171-0.186-0.483-0.284-0.938h-0.037c-0.046,0.215-0.154,0.541-0.324,0.979L13.65,31H9.643l5.074-7.539L10.009,17h4.085l2.003,4.218c0.12,0.264,0.243,0.597,0.371,1.002h0.037c0.053-0.181,0.165-0.518,0.336-1.012L18.841,17h3.805l-4.521,6.866L23.111,31H19.129z"/></svg>'
                        href = f'<a href="data:{mime_type};base64,{b64}" download="{out_filename}" style="display: flex; align-items: center; justify-content: center; background-color: #FF4B4B; color: white; padding: 0.5rem 1rem; border-radius: 0.5rem; text-decoration: none; font-weight: bold; width: 100%; box-sizing: border-box;">{svg_icon} 完成したExcelをダウンロード</a>'
                        st.markdown(href, unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"書き込みエラー: {e}")
            
            elif st.session_state.get("success_status") == "failed":
                is_timeout = st.session_state.get("is_timeout", False)
                if is_timeout:
                    st.error("⚠️ 計算がタイムアウト（30秒経過）しました。制約が複雑すぎるか、解が存在しない可能性があります。")
                else:
                    st.error("⚠️ 条件を満たす時間割が見つかりませんでした（解なし）。")
                
                diagnostics = st.session_state.get("diagnostics", [])
                if diagnostics:
                    st.warning("🔍 **原因の可能性がある設定（以下の矛盾が見つかりました）:**")
                    for d in diagnostics:
                        st.write(f"- {d}")
                else:
                    st.info("💡 単純な時数オーバーの矛盾は見つかりませんでしたが、教員同士やクラスの重複、TT（複数担当）のスケジュール調整がパズルとして解けなかったようです。「1」の枠（配置可能コマ）を増やすか、時数を減らして再度お試しください。")
                    
        except Exception as e:
            st.error(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    main()
