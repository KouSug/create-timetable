import pandas as pd
import openpyxl

# Create excel with empty column A, data in B and C
wb = openpyxl.Workbook()
ws = wb.active
ws['B1'] = 'Header1'
ws['C1'] = 'Header2'
ws['B2'] = 10
ws['C2'] = 20
wb.save("test_empty_col.xlsx")

df = pd.read_excel("test_empty_col.xlsx", engine="openpyxl")
print(df.columns)
